from flask import Flask, jsonify, request, abort
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Excel file to store car data
EXCEL_FILE = "cars.xlsx"

# Token for authentication (in a real app, use a proper authentication system)
AUTH_TOKEN = "secret_token"

# Helper function to load the Excel file
def load_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cars"
        ws.append(["ID", "Make", "Model", "Year", "Color"])
        wb.save(EXCEL_FILE)
    return load_workbook(EXCEL_FILE)

# Helper function to save the Excel file
def save_excel(wb):
    wb.save(EXCEL_FILE)

# Helper function to validate token
def validate_token():
    token = request.headers.get("Authorization")
    if token != AUTH_TOKEN:
        abort(401, description="Unauthorized: Invalid token")

# CRUD Operations

# Create a new car
@app.route("/cars", methods=["POST"])
def create_car():
    validate_token()
    data = request.json
    if not data or not all(key in data for key in ["make", "model", "year", "color"]):
        abort(400, description="Invalid input: Ensure 'make', 'model', 'year', and 'color' are provided.")

    wb = load_excel()
    ws = wb.active
    new_id = ws.max_row  # Auto-generate ID
    ws.append([new_id, data["make"], data["model"], data["year"], data["color"]])
    save_excel(wb)

    return jsonify({"message": "Car created successfully", "id": new_id}), 201

# Read all cars
@app.route("/cars", methods=["GET"])
def get_cars():
    validate_token()
    wb = load_excel()
    ws = wb.active
    cars = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        cars.append({"id": row[0], "make": row[1], "model": row[2], "year": row[3], "color": row[4]})
    return jsonify(cars)

# Read a specific car by ID
@app.route("/cars/<int:car_id>", methods=["GET"])
def get_car(car_id):
    validate_token()
    wb = load_excel()
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[0] == car_id:
            return jsonify({"id": row[0], "make": row[1], "model": row[2], "year": row[3], "color": row[4]})
    abort(404, description="Car not found")

# Update a car by ID
@app.route("/cars/<int:car_id>", methods=["PUT"])
def update_car(car_id):
    validate_token()
    data = request.json
    if not data or not any(key in data for key in ["make", "model", "year", "color"]):
        abort(400, description="Invalid input: Provide at least one field to update ('make', 'model', 'year', or 'color').")

    wb = load_excel()
    ws = wb.active
    for row in ws.iter_rows(min_row=2):  # Skip header row
        if row[0].value == car_id:
            if "make" in data:
                row[1].value = data["make"]
            if "model" in data:
                row[2].value = data["model"]
            if "year" in data:
                row[3].value = data["year"]
            if "color" in data:
                row[4].value = data["color"]
            save_excel(wb)
            return jsonify({"message": "Car updated successfully"})
    abort(404, description="Car not found")

# Delete a car by ID
@app.route("/cars/<int:car_id>", methods=["DELETE"])
def delete_car(car_id):
    validate_token()
    wb = load_excel()
    ws = wb.active
    for row in ws.iter_rows(min_row=2):  # Skip header row
        if row[0].value == car_id:
            ws.delete_rows(row[0].row)
            save_excel(wb)
            return jsonify({"message": "Car deleted successfully"})
    abort(404, description="Car not found")

# Error Handlers
@app.errorhandler(400)
def bad_request(error):
    return jsonify({"error": str(error)}), 400

@app.errorhandler(401)
def unauthorized(error):
    return jsonify({"error": str(error)}), 401

@app.errorhandler(404)
def not_found(error):
    return jsonify({"error": str(error)}), 404

# Run the server
if __name__ == "__main__":
    app.run(debug=True)